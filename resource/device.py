import openpyxl
import werkzeug
import requests # excel backup
from flask_jwt_extended import jwt_required
from flask_restful import Resource, reqparse, request
from config.properties import *
from models.device import DeviceModel, DeviceGroupModel, DeviceShotModel, DeviceLogModel
from models.user import UserModel
# from models.layer import LayerModel
# from models.device_shot import DeviceShotModel
# from models.device_usage import DeviceUsageModel
# from models.statistics_01 import Statistics01Model   0824 code delete
from werkzeug.datastructures import ContentSecurityPolicy, FileStorage
from utils.jsonutil import AlchemyEncoder as jsonEncoder
from utils.fileutil import FileUtils
from flask_login import current_user
from openpyxl    import load_workbook, Workbook
from datetime import date, datetime
import json
import pika
import logging
import random
import string
import sys
from flask import send_from_directory

# 셋톱(조회/등록/수정/삭제)
class Settop(Resource):
    parse = reqparse.RequestParser()

    parse.add_argument('dev_nm', type=str)
    parse.add_argument('dev_id', type=str)
    parse.add_argument('dev_cmt', type=str)
    parse.add_argument('device_type', type=str)
    parse.add_argument('device_location', type=str)
    parse.add_argument('device_longitude', type=str)
    parse.add_argument('device_latitude', type=str)
    parse.add_argument('device_disk_total', type=str)
    parse.add_argument('device_mem_total', type=str)
    parse.add_argument('device_ncps', type=str)
    parse.add_argument('device_group_id_0', type=str)
    parse.add_argument('device_group_id_1', type=str)
    parse.add_argument('device_group_id_2', type=str)
    parse.add_argument('user_id', type=str)

    # 셋톱 상세 조회
    def get(self, dev_id):
        settop_obj = DeviceModel.find_by_id(dev_id)
        mdfy_dt = settop_obj.mdfy_dt.strftime("%Y-%m-%d %H:%M")
        settop_obj.mdfy_dt = mdfy_dt

        # 셋톱 상세 조회 GROUPID = NULL 값, 예외처리 (JUN, 211026)
        if(settop_obj.device_group_id == None) :
            final_group_id = "0"
            settop_obj.device_group_id = final_group_id

        else :   
            final_group_id = str(settop_obj.device_group_id)
            group_obj = DeviceGroupModel.find_by_id(final_group_id)
            group_detail = group_obj.device_group_up_cd

            if(group_detail != 0):
                up_group_obj = DeviceGroupModel.find_by_id(group_detail)
                final_group_id = str(final_group_id) + "," + str(up_group_obj.device_group_id)
                up_group_detail = up_group_obj.device_group_up_cd
                if(up_group_detail != 0):
                    root_group_obj = DeviceGroupModel.find_by_id(up_group_detail)
                    final_group_id = str(final_group_id) + "," + str(root_group_obj.device_group_id)
                    settop_obj.device_group_id = final_group_id
                elif(up_group_detail == 0):
                    settop_obj.device_group_id = final_group_id
            elif(group_detail == 0):
                settop_obj.device_group_id = final_group_id
            

        settop_obj = json.dumps(settop_obj, cls=jsonEncoder)
        print(settop_obj)
        # scr_obj = [dict(r) for r in Statistics01Model.get_screen_thum(dev_id)]

        # return {'resultCode': '0', "resultString": "SUCCESS", "resultValue": settop_obj, "data": scr_obj}, 200
        return {'resultCode': '0', "resultString": "SUCCESS", "resultValue": settop_obj}, 200
    
    # 셋톱 등록
    def post(self):

        print("RECEIVED !!  Settop post !!")

        params = Settop.parse.parse_args()

        print(params)

        # dev_id = Settop.random_set_top_id()
        dev_id =                params['dev_id']
        dev_nm =                params['dev_nm']
        dev_cmt =               params['dev_cmt']
        device_type =           params['device_type']
        device_location =       params['device_location']
        device_longitude =      params['device_longitude']
        device_latitude =       params['device_latitude']
        device_disk_total =     params['device_disk_total']
        device_mem_total =      params['device_mem_total']
        device_ncps =           params['device_ncps']
        user_id =               params['user_id']
        device_group_id_0 =     params['device_group_id_0']
        device_group_id_1 =     params['device_group_id_1']
        device_group_id_2 =     params['device_group_id_2']
        rgt_dt =                datetime.now()
        mdfy_dt =               datetime.now()


        dev_id = dev_id.replace(" ", "_")

        duplicateCheck = [dict(r) for r in DeviceModel.find_by_USE_id(dev_id)]
        print(duplicateCheck)

        if(duplicateCheck == []):
            if(device_group_id_2 != "all"): 
                device_group_id = device_group_id_2
            elif(device_group_id_1 != "all"):
                device_group_id = device_group_id_1
            elif(device_group_id_0 != "all"):
                device_group_id = device_group_id_0
            elif(device_group_id_0 == "all"):
                device_group_id = "0"
            
            # device_group_id : null값 방지! (JUN)
            if(device_group_id == None) :
                device_group_id = "0"
            
            
            device_obj = DeviceModel(dev_id, dev_nm, dev_cmt, device_group_id, device_type, device_location,
                                        device_longitude , device_latitude , device_disk_total, device_mem_total, device_ncps, user_id, rgt_dt, mdfy_dt)

            try:
                # DB 저장
                
                device_obj.save_to_db()
                
                # 셋톱 갯수 업데이트 (6/22 김동수)
                # 1. user_gr확인
                user_gr = current_user.user_gr
                print("현재 로그인 한 id = " + user_id + user_gr)

                # 2. user_gr과 user_id 넣고 업데이트 하기
                resultString = UserModel.update_now_settop(UserModel, user_id, user_gr)
                
                # 3. 관리자가 아닌 경우, 하위 유저가 업로드할 경우, 상위 유저 now_disk도 업데이트
                if user_gr != '0102':
                    create_user_obj = UserModel.get_create_user_id(user_id)
                    create_user_id_dict = [dict(r) for r in create_user_obj]
                    user_id = create_user_id_dict[0]['create_user_id']
                    resultString = UserModel.update_now_settop(UserModel, user_id, '0102')

                print("user테이블 용량 업데이트 결과값: " + resultString)


            except Exception as e:

                device_obj.session_rollback()
                logging.fatal(e, exc_info=True)
                return {'resultCode': '100', "resultString": "FAIL"}, 500

            return {'resultCode': '0', "resultString": dev_nm + " 셋톱박스가 저장 되었습니다."}, 200
        
        else :

            return {'resultCode': '0', "resultString": dev_id + " 셋탑아이디가 중복 되었습니다."}, 500


    def put(self, dev_id):

        params = Settop.parse.parse_args()

        dev_nm =                params['dev_nm']
        dev_cmt =               params['dev_cmt']
        device_type =           params['device_type']
        device_location =       params['device_location']
        device_longitude =      params['device_longitude']
        device_latitude =       params['device_latitude']
        device_disk_total =     params['device_disk_total']
        device_mem_total =      params['device_mem_total']
        device_ncps =           params['device_ncps']
        device_group_id_0 =       params['device_group_id_0']
        device_group_id_1 =       params['device_group_id_1']
        device_group_id_2 =       params['device_group_id_2']
        


        print(params)
        # print(set_horizon)
        if(device_group_id_2 != "all"): 
            device_group_id = device_group_id_2
        elif(device_group_id_1 != "all"):
            device_group_id = device_group_id_1
        elif(device_group_id_0 != "all"):
            device_group_id = device_group_id_0
        elif(device_group_id_0 == "all"):
            device_group_id = "0"

        try:

            set_top_obj = DeviceModel.find_by_id(dev_id)

            set_top_obj.dev_nm =            dev_nm
            set_top_obj.dev_cmt =           dev_cmt
            set_top_obj.device_type =       device_type
            set_top_obj.device_location =   device_location
            set_top_obj.device_longitude =  device_longitude
            set_top_obj.device_latitude =   device_latitude
            set_top_obj.device_disk_total = device_disk_total
            set_top_obj.device_mem_total =  device_mem_total
            set_top_obj.device_ncps =       device_ncps
            set_top_obj.device_group_id =   device_group_id
            set_top_obj.mdfy_dt = datetime.now()

            set_top_obj.save_to_db()

        except Exception as e:
            logging.fatal(e, exc_info=True)
            return {'resultCode': '100', "resultString": "셋톱 데이터 수정에 실패하였습니다."}, 500

        return {'resultCode': '0', "resultString": dev_nm + " 데이터가 수정되었습니다."}, 200

    def delete(self,dev_id):

        print("ENTERED Device DELETE")
        print("Device ID : ["+str(dev_id)+"]")

        try:

            DeviceModel.update_device_use(dev_id)

             # 셋톱 갯수 업데이트 (6/22 이준)
            # 1. user_gr확인
            user_gr = current_user.user_gr
            user_id = current_user.user_id
            
            print("현재 로그인 한 id = " + user_id + user_gr)

            # 2. user_gr과 user_id 넣고 업데이트 하기
            resultString = UserModel.update_now_settop(UserModel, user_id, user_gr)
            
            # 3. 관리자가 아닌 경우, 하위 유저가 업로드할 경우, 상위 유저 now_settop도 업데이트
            if user_gr != '0102':
                create_user_obj = UserModel.get_create_user_id(user_id)
                create_user_id_dict = [dict(r) for r in create_user_obj]
                user_id = create_user_id_dict[0]['create_user_id']
                resultString = UserModel.update_now_settop(UserModel, user_id, '0102')

        except Exception as e:

            logging.fatal(e, exc_info=True)
            return {'resultCode': '100', "resultString": "FAIL"}, 500

        return {'resultCode': '0', "resultString": str(dev_id) + " 셋탑이 삭제 되었습니다."}, 200

    @staticmethod
    def random_set_top_id(length=40):

        letters = string.ascii_letters

        set_top_id = ''.join(random.choice(letters) for i in range(length))

        if DeviceModel.find_by_id(set_top_id):
            Settop.random_set_top_id()

        return set_top_id


# 셋톱 리스트 전체 및 검색 조회
class SettopSearch(Resource):
    parse = reqparse.RequestParser()
    parse.add_argument('parking_seq', type=int)
    parse.add_argument('parking_nm', type=str)
    parse.add_argument('device_group_id', type=str)
    parse.add_argument('dev_nm', type=str)
    parse.add_argument('dev_id', type=str)
    parse.add_argument('dev_cmt', type=str)
    parse.add_argument('dev_ethe_mac', type=str)
    # parse.add_argument('tab_gbn', type=str)

    parse.add_argument('start', type=int)
    parse.add_argument('length', type=int)


    # 셋톱 상세 조회
    def get(self):
        cnt_obj = [dict(r) for r in DeviceModel.group_device_cnt()]

        print(len(cnt_obj))

        print(type(cnt_obj))

        

        for idx in cnt_obj:
            # if Exist delivered subTitles

            print(idx)
            print(type(idx))
            
        
        # scr_obj = [dict(r) for r in Statistics01Model.get_screen_thum(dev_id)]

        # return {'resultCode': '0', "resultString": "SUCCESS", "resultValue": settop_obj, "data": scr_obj}, 200
        return {'resultCode': '0', "resultString": "SUCCESS", "resultValue": cnt_obj}, 200


    def post(self):
        res_data = {}

        params = SettopSearch.parse.parse_args()

        device_group_id = params['device_group_id']
        dev_nm = params['dev_nm']
        dev_id = params['dev_id']
        dev_cmt = params['dev_cmt']
        dev_ethe_mac = params['dev_ethe_mac']
        # tab_gbn = params['tab_gbn']
        start = params['start']
        length = params['length']
        user_gr = current_user.user_gr

        # if(current_user.user_gr == '0103'): #담당자인 경우
        # else:

        # 사용자 Total Count
        param = (device_group_id, dev_nm, dev_id, dev_cmt, dev_ethe_mac, user_gr)
        tot_list = [dict(r) for r in DeviceModel.get_set_top_count(param)]
        res_data['recordsTotal'] = tot_list[0]['tot_cnt']
        res_data['recordsFiltered'] = tot_list[0]['tot_cnt']
        # print(res_data)
        ###################################################################
        param = (device_group_id, dev_nm, dev_id, dev_cmt, dev_ethe_mac, start, length, user_gr)

        device_list = DeviceModel.get_set_top_list(param)



        final_list = [{
            'row_cnt': row[0],
            'dev_nm': row[1],
            'dev_id': row[2],
            'dev_cmt': row[3],
            'device_location': row[4],
            'device_longitude': row[5],
            'device_latitude': row[6],
            'device_disk':row[7],
            'device_cpu':row[8],
            'device_mem_used' : row[9],
            'device_sw_ver' : row[10],
            'device_play_status': row[11],
            'device_conn': row[12],
            'organ_id' : row[13],
            'device_control' : row[14],
            'parking_id': row[15],
            'user_id': row[16],
            'device_type' : row[18]
        } for row in device_list]

        res_data['data'] = final_list


        # print("final_list = ["+str(final_list)+"]")

        # 응답 결과
        res_data['resultCode'] = "0"
        res_data['resultString'] = "SUCCESS"

        return res_data, 200


class SettopMulti(Resource):
    parse = reqparse.RequestParser()

    parse.add_argument('files', type=str)

    def post(self):

        print("POST ENTERED !!  SettopMulti")

        device_list = []
        duplicated_list = []
        duplicated_count = 0
        device_reg_count = 0

        try:
            # 실제 파일 로컬에서 가져오기
            file = request.files.get("files")

            # 엑셀 파일 불러오기
            wb = load_workbook(file, data_only=True)

            # 엑셀 파일 시트 불러오기
            sheet = wb['device']

            # 첫번째 열의 모든 데이터 가져오기
            excel_count = -1
            all_values = []
            for row in sheet.rows:
                row_value = []
                for cell in row:
                    row_value.append(cell.value)                   
                all_values.append(row_value)

                excel_count += 1

            settop_total = current_user.user_settop
            settop_now = current_user.now_settop

            settop_limit = int(settop_total) - int(settop_now)

            if(settop_limit < excel_count):
                return {'resultCode': '100', "resultString": "엑셀 데이타가 등록가능 대수 보다 많습니다."}, 500
            else:
                for device in all_values:
                    # 인덱스가 0일경우 통과
                    if device[0] == '아이디(ID)':
                        continue
                    else:

                        # Device ID 빈 칸 체크
                        check_space = device[0]
                        check_space = check_space.replace(" ", "_")
                        
                        # Device ID 중복체크
                        reg_device_obj = DeviceModel.find_by_id(check_space)
                        if(reg_device_obj):
                            # 중복
                            duplicated_list.append(check_space)
                            duplicated_count += 1

                        else:

                            if check_space:
                                dev_id =            check_space
                                dev_nm =            device[1]
                                dev_cmt =           device[2]
                                device_group_id =   int(device[7])
                                device_type =       device[6]
                                device_location =   device[3]
                                device_longitude =  device[4]
                                device_latitude =   device[5]
                                device_mem_total =  '0'
                                device_disk_total = '0'
                                device_ncps =       None
                                user_id =           current_user.user_id
                                rgt_dt =            datetime.now()
                                mdfy_dt =           datetime.now()

                                # dev_id, dev_nm, dev_cmt, device_group_id, device_type, device_location,
                                # device_longitude , device_latitude , device_disk_total, device_mem_total, device_ncps, user_id, rgt_dt, mdfy_dt

                                device_obj = DeviceModel(dev_id, dev_nm, dev_cmt, device_group_id, device_type, device_location,
                                                        device_longitude , device_latitude , device_disk_total, device_mem_total, device_ncps, user_id, rgt_dt, mdfy_dt)

                                # print(device_obj)

                                device_list.append(device_obj)
                                device_reg_count += 1

                if len(device_list) > 0:
                    DeviceModel.bulk_insert(device_list)

                    # 사용자 정보에 등록 대수 update
                    # update_result = UserModel.update_user_settop_usage(user_id, device_reg_count)
                    
                    update_result = UserModel.update_now_settop(UserModel, user_id, current_user.user_gr)

                    if(update_result):
                        print("사용자 정보 UPDATE 성공 하였습니다.")
                    else: 
                        print("사용자 정보 UPDATE 실패 !!")

        except Exception as e:
            print(e)
            return {'resultCode': '100', "resultString": "FAIL"}, 500

        return {
                   'resultCode': '0',
                   "resultString": "엑셀업로드 완료. 등록: ["+str(device_reg_count)+"]건 중복제외 : ["+str(duplicated_count)+"]건"
               }, 200

# excel backup (0822 J)
class ExcelDown(Resource):

    def get(self):
        

        book = Workbook()
        sheet = book.active

        sheet.title = "device"

        now = datetime.now()
        user_id = current_user.user_id

        file_name = "SettopList_"+now.strftime("%Y-%m-%d_%H:%M:%S")+".xlsx"

        settop_list = [dict(r) for r in DeviceModel.excel_set_top_list(user_id)]

        if len(settop_list) > 0:

            sheet.cell(row=1, column=1).value = "아이디(ID)"
            sheet.cell(row=1, column=2).value = "이름(NAME)"
            sheet.cell(row=1, column=3).value = "설명(COMMENT)"
            sheet.cell(row=1, column=4).value = "주소(ADDRESS)"
            sheet.cell(row=1, column=5).value = "경도(LONGITUDE)"
            sheet.cell(row=1, column=6).value = "위도(LATITUDE)"
            sheet.cell(row=1, column=7).value = "장비타입"
            sheet.cell(row=1, column=8).value = "그룹아이디"

            for idx in range(len(settop_list)):

                sheet.cell(row=idx+2, column=1).value = settop_list[idx]['dev_id']
                sheet.cell(row=idx+2, column=2).value = settop_list[idx]['dev_nm']
                sheet.cell(row=idx+2, column=3).value = settop_list[idx]['dev_cmt']
                sheet.cell(row=idx+2, column=4).value = settop_list[idx]['device_location']
                sheet.cell(row=idx+2, column=5).value = settop_list[idx]['device_longitude']
                sheet.cell(row=idx+2, column=6).value = settop_list[idx]['device_latitude']
                sheet.cell(row=idx+2, column=7).value = settop_list[idx]['device_type']
                sheet.cell(row=idx+2, column=8).value = settop_list[idx]['device_group_id']


            book.save(filename = excel_file + file_name)

            # 다운로드 처리

            dls = excel_url+file_name
            resp = requests.get(dls)

            output = open(excel_file+file_name, 'rb')
            return send_from_directory(excel_file, file_name)


class SettopInit(Resource):
    parse = reqparse.RequestParser()

    def put(self, dev_id):

        try:
            set_top_obj = DeviceModel.find_by_id(dev_id)
            set_top_obj.parking_seq = None
            set_top_obj.dev_nm = None
            set_top_obj.dev_cmt = None
            set_top_obj.dev_ip = None
            set_top_obj.mdfy_dt = datetime.now()

            set_top_obj.save_to_db()

        except Exception as e:
            logging.fatal(e, exc_info=True)
            return {'resultCode': '100', "resultString": "셋톱 초기화에 실패하였습니다.."}, 500

        return {'resultCode': '0', "resultString": "셋톱이 초기화 되었습니다."}, 200


class SettopCommSearch(Resource):

    print("Entered SettopCommSearch")

    parse = reqparse.RequestParser()
# 삭제 확인 (by. jun)
    # parse.add_argument('parking_seq', type=int)
    parse.add_argument('group_seq', type=int)

    def get(self):

        res_data = {}

        try:
            user_id = current_user.user_id
            group_list = DeviceGroupModel.get_group_list_user_id(user_id)
            final_group_list = [{
                'id': row[0],
                'pId': row[1],
                'name' : row[2]
            } for row in group_list]

            res_data['data'] = final_group_list  

        except Exception as e:
            logging.fatal(e, exc_info=True)
            return {'resultCode': '100', "resultString": "셋톱 정보 조회에 실패하였습니다."}, 500
        return {'resultCode': '0', "resultString": "셋톱 조회에 성공하였습니다.", "data": res_data}, 200


class SettopSimpleSearch(Resource):
    print("Entered SettopSimpleSearch")

    parse = reqparse.RequestParser()

    def get(self):

        res_data = {}

        try:
            user_id = current_user.user_id

            device_list = DeviceModel.get_select_device_list(user_id)

            final_list = [{
                'dev_id' : row[0],
                'dev_nm' : row[1],
                'dev_cmt': row[2],
                'dev_group_id' : row[3]
            } for row in device_list]

            res_data['data'] = final_list

        except Exception as e:
            logging.fatal(e, exc_info=True)
            return {'resultCode': '100', "resultString": "셋톱 정보 조회에 실패하였습니다."}, 500
        return {'resultCode': '0', "resultString": "셋톱 조회에 성공하였습니다.", "data": res_data}, 200

class SettopPush(Resource):
    # Device connection
    def __init__(self):
        self.url = 'gstech.iptime.org'
        self.port = 5672
        self.vhost = 'did'
        self.cred = pika.PlainCredentials('cms', '12345')
        self.exchange_name = 'EXC_COMMAND'
        # return

    print(("Entered SettopPush"))

    parse = reqparse.RequestParser()

    performed = False
    # shot_no = '0'

    def device_push(self,dev_id, command, schedule_id, volume, comment, duration):

        shot_no = '0'

        # paramm check
        print("RECEIVED ID :            ["+dev_id+"]")
        print("RECEIVED COMMAND :       ["+command+"]")
        print("RECEIVED SCHEDULE_ID :   ["+str(schedule_id)+"]")
        print("RECEIVED VOLUME :        ["+str(volume)+"]")
        print("RECEIVED comment :       ["+comment+"]")
        print("RECEIVED duration :      ["+str(duration)+"]")

        try:
            conn = pika.BlockingConnection(pika.ConnectionParameters(self.url, self.port, self.vhost, self.cred))
            channel = conn.channel()

            # check step 1
            # print("connection result")
            # print(conn)
            # print(channel)
            user_id = current_user.user_id

            if(command == 'schedule'):
                command_object = {
                    "command_type": command,
                    "schedule_id" : schedule_id,
                    "volume": 0,
                    "emergency_text": "",
                    "emergency_duration": 0
                }
            elif(command == 'status'):
                # 전체 device list에 시스템 명령 전달 (시스템 데이타 올리기)

                device_list = ""
                count = 0

                dev_list = [dict(r) for r in DeviceModel.get_settop_list_by_id(user_id)]

                for r in dev_list:
                    if(device_list == ''): 
                        device_list = r["dev_id"]
                    else:
                        device_list += "."+r["dev_id"]

                    count += 1

                    # check size of total device id list 

                    if(sys.getsizeof(device_list) > 230):
                        # 안전상 230에서 자름
                        print("Entered safe mode sending now .. ")

                        routing_key = device_list

                        command_object = {
                            "command_type": 'status',
                            "volume": 0,
                            "emergency_text": "",
                            "emergency_duration": 0
                        }

                        try:
                            channel.basic_publish(
                                exchange=self.exchange_name,
                                routing_key=routing_key,
                                body=json.dumps(command_object)
                            )
                        except Exception as e:
                            conn.close()
                            return False                   
                        device_list = ""


                print("USER ID :["+user_id+"] COUNT :["+str(count)+"] SIZE :["+str(sys.getsizeof(device_list))+"]")
                print("DEVICE LIST : ["+device_list+"]")

                # var url = "/settop/push?dev_id="+device_id+"&command=schedule&schedule_id="+schedule_id;
                # http://218.38.232.231:7077/settop/push?dev_id=&command=device&schedule_id=

                dev_id = device_list

                command_object = {
                    "command_type": 'status',
                    "volume": 0,
                    "emergency_text": "",
                    "emergency_duration": 0
                }

            elif(command == 'screen_shot'):
                # Screen shot 기능. DB 생성 기능을 주고 결과값으로 
                

                # def __init__(self, dev_id, shot_nm, shot_url, shot_thu_url, is_view):
                nowTime = datetime.now()
                shotObj = DeviceShotModel(dev_id,'','','','Y', nowTime, nowTime)
                shotObj.save_to_db()
                shot_no = shotObj.shot_no

                command_object = {
                    "command_type": command,
                    "volume": volume,
                    "emergency_text": comment,
                    "emergency_duration": int(duration)
                }
            elif(command == 'log'):

                # def __init__(self, dev_id, shot_nm, shot_url, shot_thu_url, is_view):
                nowTime = datetime.now()
                logObj = DeviceLogModel(dev_id,'','','','Y', nowTime, nowTime)
                logObj.save_to_db()
                log_no = logObj.log_no

                command_object = {
                    "command_type": command,
                    "volume": volume,
                    "emergency_text": comment,
                    "emergency_duration": int(duration)
                }

            else:
                command_object = {
                    "command_type": command,
                    "volume": volume,
                    "emergency_text": comment,
                    "emergency_duration": int(duration)
                }
                # Debug general values
                print("command :["+command+"]")
                print("volume :["+str(volume)+"]")
                print("comment :["+comment+"]")
                print("duration :["+str(duration)+"]")

            channel.exchange_declare(exchange=self.exchange_name, exchange_type='topic')
            print(channel)

            # routing_key = 'stb01.stb02.stb03'   # 전송할 셋탑박스 아이디를 dot(.)으로 구분하여 연결(최대 255byte)
            routing_key = dev_id

            try:
                channel.basic_publish(
                    exchange=self.exchange_name,
                    routing_key=routing_key,
                    body=json.dumps(command_object)
                )
            except Exception as e:
                conn.close()
                return False
            
            conn.close()



            # Device 제어 이력 Update
            # device id = 'A-1'.'B-2'.C-5' 의 리스트로 넘어 옴. 
            device_id_list = dev_id.split('.')

            # print(device_id_list)
            for dev_id_one in device_id_list:
                try:
                    # print("EACH DEVICE ID : ["+dev_id_one+"]")
                    set_top_obj = DeviceModel.find_by_id(dev_id_one)

                    set_top_obj.device_control =    command
                    set_top_obj.mdfy_dt =           datetime.now()

                    set_top_obj.save_to_db()
                except Exception as e:
                    print("EXCEPTION OCCURRED !!!!! HERE error : ["+str(e)+"]")

            print("All done //// FINE")
            # return True

            if(command == "screen_shot"):
                return shot_no
            elif(command == "log"):
                return log_no
            else:
                return True
        # connection error
        except Exception as e:
            return False

    def get(self):
        
        # Get entered obtain device id and 
        # global shot_no

        dev_id  =       request.args.get("dev_id")
        command =       request.args.get("command")
        volume =        request.args.get("volume")
        schedule_id =   request.args.get("schedule_id")
        comment =       request.args.get("comment")
        duration =      request.args.get("duration")

        if(schedule_id == None):
            schedule_id = 0

        if(duration == None):
            duration = 0

        if(volume == None):
            volume = 0

        if(comment == None):
            comment = ""

        # Debug PUSH input data 
        print("dev_id :["+dev_id+"]")
        print("command :["+command+"]")
        print("volume :["+str(volume)+"]")
        print("schedule_id :["+str(schedule_id)+"]")
        print("comment :["+comment+"]")
        print("duration :["+str(duration)+"]")

        Result = SettopPush.device_push(self, dev_id, command, schedule_id, volume, comment, duration)

        print(Result)

        if(Result):
            if(command == "screen_shot"):
                return {'resultCode': '0', 'shot_no': Result ,"resultString": "screen shot PUSH에 성공 했습니다.."}, 200
            elif(command == "log"):
                return {'resultCode': '0', 'log_no': Result ,"resultString": "log PUSH에 성공 했습니다.."}, 200
            else:
                 return {'resultCode': '0', "resultString": "단말 PUSH에 성공 했습니다.."}, 200
        else:
            return {'resultCode': '100', "resultString": "셋톱 PUSH에 실패 하였습니다."}, 200


class SettopTree(Resource):

    # print("Entered SettopTree")

    parse = reqparse.RequestParser()

    parse.add_argument('group_name', type=str)
    parse.add_argument('tID', type=str)
    parse.add_argument('depth', type=str)

    def get(self):

        # print("Entered GET SettopTree")

        tID = request.args.get("tID")
        tNAME = request.args.get("tNAME")
        tPID = request.args.get("tPID")
        depth = request.args.get("depth")

        # print("tID :    ["+tID+"]")
        # print("tNAME :  ["+tNAME+"]")
        # print("tPID :   ["+tPID+"]")


        group_obj = DeviceGroupModel.find_by_id(tID)

        if(group_obj):

            print("Data Obtain OK")

            group_obj.device_group_nm = tNAME
            group_obj.device_group_up_cd = tPID
            group_obj.device_group_depth = depth
            try:
                group_obj.save_to_db()
                print("그룹정보 변경 성공 ...")
                return {'resultCode': '0', "resultString": "그룹정보 변경에 성공 하였습니다."}, 200

            except Exception as e:
                print("그룹정보 변경 실패 ...")
                logging.fatal(e, exc_info=True)
                return {'resultCode': '100', "resultString": "그룹정보 변경에 실패하였습니다."}, 200

        return {'resultCode': '100', "resultString": "그룹이 존재하지 않습니다.."}, 200
    
    def post(self):
        # print("Entered POST SettopTree !!")

        params = SettopTree.parse.parse_args()

        print(params)

        user_id = current_user.user_id

        # dev_id = Settop.random_set_top_id()
        group_name =            params['group_name']
        tID =                   params['tID']
        depth =                 params['depth']
        rgt_dt =                datetime.now()
        mdfy_dt =               datetime.now()

        #  device_group_nm, device_group_up_cd, user_id, rgt_dt, mdfy_dt

        group_obj = DeviceGroupModel(group_name, tID, depth, user_id, rgt_dt, mdfy_dt)

        try:
            # DB 저장
            group_obj.save_to_db()

            # 저장된 control_id Obtain
            device_group_id = group_obj.device_group_id

        except Exception as e:
            logging.fatal(e, exc_info=True)
            return {'resultCode': '100', "resultString": "FAIL"}, 500

        return {'resultCode': '0', "resultString": group_name + " 그룹이 저장 되었습니다.", "device_group_id": str(device_group_id), "group_name": group_name}, 200
    
    def delete(self,tID):

        # print("Entered delete SettopTree")

        # print("tID :    ["+tID+"]")

        try:
            group_obj = DeviceGroupModel.find_by_id(tID)
            group_obj.delete_to_db()

        except Exception as e:
            logging.fatal(e, exc_info=True)
            return {'resultCode': '100', "resultString": "FAIL"}, 500

        return {'resultCode': '0', "resultString": "그룹 트리가 삭제 되었습니다."}, 200


class Settopallcnt(Resource):
    parse = reqparse.RequestParser()

    def get(self):

        res_data = {}

        try:
            user_id = current_user.user_id
            group_list = DeviceModel.cnt_device_status()
            final_group_list = [{
                'date': row[0],
                'conut': row[1],
    
            } for row in group_list]

            res_data['data'] = final_group_list  

        except Exception as e:
            logging.fatal(e, exc_info=True)
            return {'resultCode': '100', "resultString": "셋톱 정보 조회에 실패하였습니다."}, 500
        return {'resultCode': '0', "resultString": "셋톱 조회에 성공하였습니다.", "data": res_data}, 200
        
# mysql> desc tbl_device_shot;
# +--------------+-------------+------+-----+-------------------+-------------------+
# | Field        | Type        | Null | Key | Default           | Extra             |
# +--------------+-------------+------+-----+-------------------+-------------------+
# | shot_no      | int         | NO   | PRI | NULL              | auto_increment    |
# | dev_id       | varchar(45) | NO   | PRI | NULL              |                   |
# | shot_nm      | varchar(45) | YES  |     | NULL              |                   |
# | shot_url     | varchar(45) | YES  |     | NULL              |                   |
# | shot_thu_url | varchar(45) | YES  |     | NULL              |                   |
# | is_view      | varchar(1)  | NO   |     | Y                 |                   |
# | upload_yn    | varchar(1)  | NO   |     | N                 |                   |
# | create_dt    | datetime    | YES  |     | CURRENT_TIMESTAMP | DEFAULT_GENERATED |
# | modify_dt    | datetime    | YES  |     | CURRENT_TIMESTAMP | DEFAULT_GENERATED |
# +--------------+-------------+------+-----+-------------------+-------------------+

class ScreenShotSearch(Resource):
    parse = reqparse.RequestParser()

    parse.add_argument('shot_no', type=str)

    def get(self):

        print("GET ENTERRED ScreenShotSearch !!!")

        shot_no = request.args.get("shot_no")

        # shot_no 로 조회 함. update 해 주어야 함. 

        try:
            screenshotObj  = DeviceShotModel.find_by_no(shot_no)

            returnValue = {
                    'resultCode' : '0',
                    'resultString' : "스크린샷 조회에 성공하였습니다.",
                    'dev_id': screenshotObj.dev_id,
                    'shot_nm': screenshotObj.shot_nm,
                    'shot_url': screenshotObj.shot_url,
                    'shot_thu_url': screenshotObj.shot_thu_url,
                    'is_view': screenshotObj.is_view,
                    'upload_yn': screenshotObj.upload_yn,
                }
        except Exception as e:
            logging.fatal(e, exc_info=True)
            return {'resultCode': '100', "resultString": "스크린샷 조회에 실패하였습니다."}, 500

        return returnValue, 200

# mysql> desc tbl_device_log;
# +-------------+-------------+------+-----+-------------------+-------------------+
# | Field       | Type        | Null | Key | Default           | Extra             |
# +-------------+-------------+------+-----+-------------------+-------------------+
# | log_no      | int         | NO   | PRI | NULL              | auto_increment    |
# | dev_id      | varchar(45) | NO   | PRI | NULL              |                   |
# | log_nm      | varchar(45) | YES  |     | NULL              |                   |
# | log_url     | varchar(45) | YES  |     | NULL              |                   |
# | log_thu_url | varchar(45) | YES  |     | NULL              |                   |
# | is_view     | varchar(1)  | NO   |     | Y                 |                   |
# | upload_yn   | varchar(1)  | NO   |     | N                 |                   |
# | create_dt   | datetime    | YES  |     | CURRENT_TIMESTAMP | DEFAULT_GENERATED |
# | modify_dt   | datetime    | YES  |     | CURRENT_TIMESTAMP | DEFAULT_GENERATED |
# +-------------+-------------+------+-----+-------------------+-------------------+
class DeviceLogSearch(Resource):
    parse = reqparse.RequestParser()

    parse.add_argument('log_no', type=str)

    def get(self):

        # print("GET ENTERRED DeviceLogSearch !!!")

        log_no = request.args.get("log_no")

        # log_no 로 조회 함. update 해 주어야 함. 

        try:
            deviceLogObj  = DeviceLogModel.find_by_no(log_no)

            returnValue = {
                    'resultCode' : '0',
                    'resultString' : "로그파일 조회에 성공하였습니다.",
                    'dev_id': deviceLogObj.dev_id,
                    'log_nm': deviceLogObj.log_nm,
                    'log_url': deviceLogObj.log_url,
                    'log_thu_url': deviceLogObj.log_thu_url,
                    'is_view': deviceLogObj.is_view,
                    'upload_yn': deviceLogObj.upload_yn,
                }
        except Exception as e:
            logging.fatal(e, exc_info=True)
            return {'resultCode': '100', "resultString": "로그파일 조회에 실패하였습니다."}, 500

        return returnValue, 200
        

class DeviceCurrentCount(Resource):

    def get(self):

        # print("GET ENTERRED DeviceCurrentCount !!!")

        try:

            current_cnt = 0
            current_total = 0
            current_fault_cnt = 0

            currentCountObj     = [dict(r) for r in DeviceModel.get_current_device_cnt()]
            currentFaultCountObj     = [dict(r) for r in DeviceModel.get_current_fault_device_cnt()]
            currentTotalObj     = [dict(r) for r in DeviceModel.get_current_device_total()]
            


            for r in currentCountObj:
                current_cnt += 1

            for r in currentFaultCountObj:
                current_fault_cnt += 1


            for r in currentTotalObj:
                current_total = r["cnt"]


            # print(currentCountObj)
            # print(currentTotalObj)

            returnValue = {
                    'resultCode' : '0',
                    'resultString' : "실시간 데이타 조회에 성공하였습니다.",
                    'current_cnt':  current_cnt,
                    'current_fault_cnt':  current_fault_cnt,
                    'current_total': current_total
                }

        except Exception as e:
            logging.fatal(e, exc_info=True)
            return {'resultCode': '100', "resultString": "실시간 데이타 조회에 실패하였습니다."}, 500

        return returnValue, 200




# 편성정보 확인
class SettopOrgan(Resource):
    parse = reqparse.RequestParser()

    parse.add_argument('start', type=int)
    parse.add_argument('length', type=int)
    parse.add_argument('dev_id', type=str)

    def post(self, dev_id):

        params = SettopOrgan.parse.parse_args()
        
        start = params['start']
        length = params['length']

        user_id =    current_user.user_id
        tot_list = [dict(r) for r in DeviceModel.organ_by_dev_id_cnt(dev_id)]

        organ_list = [dict(r) for r in DeviceModel.organ_by_dev_id(dev_id, length, start)] 
        return {
                "recordsTotal": tot_list[0]['tot_cnt'],
                "recordsFiltered": tot_list[0]['tot_cnt'],
                "resultCode": '0',
                "resultString": "SUCCESS",
                "data" : organ_list
               },  200

class DeviceDuplicate(Resource):
    parse = reqparse.RequestParser()

    def get(self, dev_id):
        settop_obj = [dict(r) for r in DeviceModel.find_by_USE_id(dev_id)]
        print(settop_obj)
        if settop_obj:
            return {'result': True}
        else:
            return {'result': False}

